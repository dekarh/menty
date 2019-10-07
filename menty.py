# -*- coding: utf-8 -*-
# Перечень ВСЕХ полей (с примером заполнения) во ВСЕХ таблицах для СБ

import sys, argparse
from _datetime import datetime, timedelta, date
import os
from mysql.connector import MySQLConnection, Error
from collections import OrderedDict
import openpyxl
from pymongo import MongoClient
import psycopg2

from lib import read_config, fine_phone

MYSQL_DBS = ['passport_service', 'saturn_crm', 'saturn_fin']

wb_rez = openpyxl.Workbook(write_only=True)
ws_rez = []

#q = """
# -------- Mongo ------------------------------------------------------
# подключаемся к серверу
cfg = read_config(filename='menty.ini', section='Mongo')
conn = MongoClient('mongodb://' + cfg['user'] + ':' + cfg['password'] + '@' + cfg['ip'] + ':' + cfg['port'] + '/'
                   + cfg['db'])
# выбираем базу данных
db = conn.saturn_v
# перебираем все коллекции документов
komb = 0
for i, coll_name in enumerate(db.collection_names()):
    ws_rez.append(wb_rez.create_sheet('Mongo->' + coll_name))
    coll = db[coll_name]
    fields_checked = []
    #product_alias: raiffeisen_loan_lead_referral
    for j, doc in enumerate(coll.find({})):
        fields = list(doc.keys())
        fields2check = ''.join(sorted(fields))
        if fields2check not in fields_checked:
            fields_checked.append(fields2check)
            komb += 1
            ws_rez[i].append([])
            ws_rez[i].append(['Комбинация № ', komb])
            ws_rez[i].append(fields)
            fields_rez = []
            for field in fields:
                if str(type(doc[field])).find('str') < 0 and str(type(doc[field])).find('int') < 0:
                    fields_rez.append(str(doc[field]))
                else:
                    fields_rez.append(doc[field])
            ws_rez[i].append(fields_rez)

# ----------- MySQL -----------------------------------------------------
cfg = read_config(filename='menty.ini', section='SaturnOPS')
dbconn = MySQLConnection(**cfg)
tables_cursor = dbconn.cursor()
for mysql_db in MYSQL_DBS:
    ws_rez.append(wb_rez.create_sheet('MySQL->' + mysql_db))
    tables_cursor.execute('SHOW TABLES IN ' + mysql_db)
    tables_rows = tables_cursor.fetchall()
    col_cursor = dbconn.cursor()
    cursor = dbconn.cursor()
    for table_row in tables_rows:
        col_cursor.execute('SELECT * FROM ' + mysql_db + '.' + table_row[0] + ' LIMIT 0,2')
        col_rows = col_cursor.fetchall()
        col_names = col_cursor.column_names
        sql = 'SELECT * FROM ' + mysql_db + '.' + table_row[0] + ' ORDER BY ' + col_names[0] + ' DESC LIMIT 0,5'
        cursor.execute(sql)
        rows = cursor.fetchall()
        ws_rez[len(ws_rez) - 1].append([])
        ws_rez[len(ws_rez) - 1].append([table_row[0]])
        ws_rez[len(ws_rez) - 1].append(col_names)
        if len(rows):
            for row in rows:
                col_rez = []
                for j, col_name in enumerate(col_names):
                    if row[j]:
                        col_rez.append(row[j])
                    else:
                        col_rez.append('--пусто--')
                ws_rez[len(ws_rez) - 1].append(col_rez)
        else:
            col_rez = []
            for j, col_name in enumerate(col_names):
                col_rez.append('--пусто--')
            ws_rez[len(ws_rez) - 1].append(col_rez)
#"""
# ----------- Postgres --------------------------------------------------
q1 = """
cfg = read_config(filename='menty.ini', section='postgresql')
conn = psycopg2.connect(**cfg)
tables_cursor = conn.cursor()
sql = "SELECT table_name FROM information_schema.tables WHERE table_schema NOT IN " \
      "('information_schema','pg_catalog') AND table_schema IN('public', 'myschema')"
tables_cursor.execute(sql)
cursor = conn.cursor()
ws_rez.append(wb_rez.create_sheet('ЧтобыРаботало'))
ws_rez.append(wb_rez.create_sheet('Postgresql'))
for table_row in tables_cursor:
    cursor.execute('SELECT * FROM ' + table_row[0] + ' LIMIT 2')
    col_names = [desc[0] for desc in cursor.description]
    cursor.execute('SELECT * FROM ' + table_row[0] + ' ORDER BY ' + col_names[0] + ' DESC LIMIT 5')
    ws_rez[len(ws_rez) - 1].append([])
    ws_rez[len(ws_rez) - 1].append([table_row[0]])
    ws_rez[len(ws_rez) - 1].append(col_names)
    cycled = False
    for row in cursor:
        col_rez = []
        cycled = True
        for j, col_name in enumerate(col_names):
            if row[j]:
                if str(type(row[j])).find('str') < 0:
                    col_rez.append(str(row[j]))
                else:
                    col_rez.append(row[j])
            else:
                col_rez.append('--пусто--')
        ws_rez[len(ws_rez) - 1].append(col_rez)
    if not cycled:
        col_rez = []
        for j, col_name in enumerate(col_names):
            col_rez.append('--пусто--')
        ws_rez[len(ws_rez) - 1].append(col_rez)

"""
wb_rez.save(datetime.now().strftime('%Y-%m-%d_%H-%M') + 'databases.xlsx')